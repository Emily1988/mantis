from xlrd import open_workbook
from xlutils.copy import copy
from mantis import date
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
import openpyxl

def sen_wk():
    wk = []
    for i in range(1,53):
        wk.append('wk%s'%i)
    return wk
#定义单元格格式
# 居中对齐
def style_alignment():
    alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False,
                          shrink_to_fit=False, indent=0)
    return alignment

def style_border():
    border = Border(left=Side(border_style=None,
        color = 'FF000000'),
        right = Side(border_style=None,
        color = 'FF000000'),
        top = Side(border_style=None,
        color = 'FF000000'),
        bottom = Side(border_style=None,
        color = 'FF000000'),
        diagonal = Side(border_style=None,
        color = 'FF000000'),
        diagonal_direction = 0,
        outline = Side(border_style=None,
        color = 'FF000000'),
        vertical = Side(border_style=None,
        color = 'FF000000'),
        horizontal = Side(border_style=None,
        color = 'FF000000'))

    return border

def write03(ws,dict,monthdict):

    for index,(key,value) in enumerate(dict.items()):
        # print(index,key,value)alignment
        ws.cell(3, index + 3).alignment = style_alignment()
        ws.cell(4, index + 3).alignment = style_alignment()  # wk提交的bug
        ws.cell(5, index + 3).alignment = style_alignment()  # wk解决的bug
        ws.cell(6, index + 3).alignment = style_alignment()  # wk+1解决的bug
        ws.cell(7, index + 3).alignment = style_alignment()  # wk+2解决的bug
        ws.cell(8, index + 3).alignment = style_alignment()  # wk+3解决的bug
        ws.cell(9, index + 3).alignment = style_alignment()  # 3周内未提供对策
        ws.cell(10, index + 3).alignment = style_alignment() #遗留未解决问题数

        ws.cell(3, index + 3).value = key
        if value[0] == 0:
            ws.cell(4, index + 3).value = '/'  # wk提交的bug
            ws.cell(5, index + 3).value = '/'  # wk解决的bug
            ws.cell(6, index + 3).value = '/'  # wk+1解决的bug
            ws.cell(7, index + 3).value = '/'  # wk+2解决的bug
            ws.cell(8, index + 3).value = '/'  # wk+3解决的bug
            ws.cell(9, index + 3).value = '/'  # 3周内未提供对策
            ws.cell(10, index + 3).value = '/'  # 3周内未提供对策
        else:
            ws.cell(4, index + 3).value = value[0]  # wk提交的bug
            ws.cell(5, index + 3).value = value[1]  # wk解决的bug
            ws.cell(6, index + 3).value = value[2]  # wk+1解决的bug
            ws.cell(7, index + 3).value = value[3]  # wk+2解决的bug
            ws.cell(8, index + 3).value = value[4]  # wk+3解决的bug
            ws.cell(9, index + 3).value = value[5]  # 3周内未提供对策
            ws.cell(10, index + 3).value = value[6]  # 遗留未解决问题数

            ws.cell(9, index + 3).font = Font(color="FF0000")  # wk+3解决的bug
            ws.cell(10, index + 3).font = Font(color="FF0000")  # 遗留未解决问题数
            if ws.cell(9, index + 3).value == 0:
                ws.cell(9, index + 3).font = Font(color="000000")  # wk+3解决的bug
            if ws.cell(10, index + 3).value == 0:
                ws.cell(10, index + 3).font = Font(color="000000")  # 遗留未解决问题数
    i = 0
    for index,(key,value) in enumerate(monthdict.items()):
        # print(index,key,value)
        first = list(value[0].keys())[0]
        last = list(value[len(value)-1].keys())[0]
        a = '0'
        b = '0'
        for row in ws['C3:BB3']:  # 返回的row是一个tuple对象
            for cell in row:
                if cell.value == first:
                    a = cell.column
                    # print('first column: %s'%cell.column)
                if cell.value == last:
                    b = cell.column
                    # print('last column: %s'%cell.column)
                # print('row: %s  column: %s  value: %s' % (cell.row, cell.column, cell.value))

        rang_cell = '%s1:%s1'%(a,b)  #C=67
        # print(rang_cell)
        # 按照月份划分
        ws.merge_cells('%s1:%s1' % (a, b))
        ws['%s1' %a] = key
        ws['%s1'%a].alignment = style_alignment()

        # 按照周划分
        for v in value:
            start = list(v.values())[0][0]
            leng = len(list(v.values())[0])
            end = list(v.values())[0][leng-1]
            dwk_period = "%s~%s" % (start, end)
            ws.cell(2,i+3).value = dwk_period
            ws.cell(2,i+3).alignment = style_alignment()
            i = i+1


def dynamic_list():
    #1. 准备locals()函数
    wk = locals()
    #2. 循环52次，从1~52赋值给locals函数对应的变量wk
    #wk是一个字典,所以将我们需要的列表名称作为key值传入
    for i in range(1,53):
        str = "wk%s"%i
        wk[str] = []
        return wk[str]

def test_dynamic_list():
    wk = locals()
    for i in range(1,53):
        str = "wk%s"%i
        wk[str] = []
        wk[str].append('我是第%s个list'%i)
    print(wk['wk1'])
    print(wk['wk2'])

#根据传入字典的周数目，和日期范围，得到wk周列表
def getwkdic_dateRange(beginDate, endDate,dict):
    # 动态创建len(dict)个数的list,（wk1=[],wk2=[],wk3=[]...）
    # 这些列表用于存放wk周内的day（2017年wk1[2017-01-02,2017-01-03,2017-01-04,2017-01-05,2017-01-06,2017-01-07,2017-01-08]）
    wk = locals()
    for i in range(1, len(dict) + 1):
        str = "wk%s" % i
        wk[str] = []

    for i, _date in enumerate(date.dateRange('2017-01-01', '2017-12-31')):
        timelist = _date.split('-')
        wklist = date.week_date(timelist[0], timelist[1], timelist[2])
        str = 'wk%s' % wklist[1]
        keylist = list(dict.keys())
        if i == 0 and str == 'wk52':
            # 2017年的头几天有是2016年的wk52周，所以要排除在外
            continue
        else:
            index = keylist.index(str)  # 找到列表种的对应元素的位置i
            wk['wk%s' % (index + 1)].append(_date)

    monthdict = {}
    wkdict = {}
    dictv = []   #存放字典的值
    for i in range(1, len(dict)+1):
        # del wk['wk%s' % i][1:len(wk['wk%s' % i])-1]
        datelist = wk['wk%s' % i]  #周列表
        end_index = len(wk['wk%s' % i])-1
        #wk?中第一个元素的年，月，日
        start_year = datelist[0].split('-')[0]
        start_month = datelist[0].split('-')[1].lstrip('0')
        start_day = datelist[0].split('-')[2].lstrip('0')
        # wk?中最后一个元素的年，月，日
        end_year = datelist[end_index].split('-')[0]
        end_month = datelist[end_index].split('-')[1].lstrip('0')
        end_day = datelist[end_index].split('-')[2].lstrip('0')
        #修改周列表中第一个元素和最后一个元素的值
        wk['wk%s' % i][0] = start_day
        wk['wk%s' % i][end_index] = end_day
        wkdict.setdefault('wk%s' % i, wk['wk%s' % i])
        if int(start_day) < int(end_day):
            if start_day == '1' and end_day == '7':   #遇月末不交替
                dictv = []
            listdic = {'wk%s' % i: wk['wk%s' % i]}
            # dictv.append('wk%s' % i)
            dictv.append(listdic)
            monthdict.setdefault('%s年%s月' %(start_year,start_month),dictv)
        else:
            dictv = []  # 存放字典的值
            # print('月份交替界限')
            for j in range(0,len(datelist)-1):
                day1 = 0
                day2 = 0
                if j == 0:
                    day1 = datelist[j]
                    day2 = datelist[j+1].split('-')[2].lstrip('0')
                elif j == len(datelist)-2:
                    day1 = datelist[j].split('-')[2].lstrip('0')
                    day2 = datelist[len(datelist)-1]
                else:
                    day1 = datelist[j].split('-')[2].lstrip('0')
                    day2 = datelist[j+1].split('-')[2].lstrip('0')
                if int(day1) > int(day2):
                    if j >= 3:
                        # print('#说明wk?就近day1')
                        # print(wk['wk%s' % i],'wk%s' % i,start_month)
                        key = '%s年%s月' % (start_year, start_month)
                        # value = 'wk%s'%i
                        value = {'wk%s' % i: wk['wk%s' % i]}
                        if key not in monthdict.keys():
                            # dictv.append('wk%s' % i)
                            monthdict[key] = dictv.append(value)
                        else:
                            monthdict[key].append(value)
                    else:  #说明wk?就近day2
                        # print('#说明wk?就近day2')
                        key = '%s年%s月' % (start_year, end_month)
                        value = {'wk%s' % i: wk['wk%s' % i]}
                        # print(wk['wk%s' % i],'wk%s' % i,end_month)
                        if key not in monthdict.keys():
                            dictv.append(value)
                            monthdict[key] = dictv
                        else:
                            monthdict[key]= dictv
    print(monthdict)
    return monthdict

#初始工作
def run_excel(wb,dict,projectname,index):
    monthdict = getwkdic_dateRange('2017-01-01', '2017-12-31', dict)
    workbook(wb,dict,monthdict,projectname,index)

def workbook(wb,dict,monthdict,projectname,index):

    wb.create_sheet(projectname,index)
    print(wb.sheetnames)
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # ws = wb.active
    ws = wb[projectname]
    ws.merge_cells('A3:A10')

    print(ws.merged_cells)
    ws['A3'] = projectname
    ws['A3'].alignment = style_alignment()
    ws['B1'] = '2017'
    ws['B4'] = '提交Maj bug\rMaj bug解决'
    ws['B5'] = 'Week'
    ws['B6'] = 'Week+1'
    ws['B7'] = 'Week+2'
    ws['B8'] = 'Week+3'
    ws['B9'] = '3周内未提供对策'
    ws['B10'] = '遗留未解决问题数'
    ws['B6'].fill = PatternFill(fill_type='solid', fgColor='FFFF00', end_color='FFFF00')
    ws['B7'].fill = PatternFill(fill_type='solid', fgColor='FFFF00', end_color='FFFF00')
    ws['B8'].fill = PatternFill(fill_type='solid', fgColor='FFFF00', end_color='FFFF00')
    ws['B9'].fill = PatternFill(fill_type='solid', fgColor='FF0000', end_color='FF0000')
    ws['B10'].fill = PatternFill(fill_type='solid', fgColor='B22222', end_color='B22222')

    write03(ws, dict, monthdict)
    wb.save('New11.xls')



if __name__ == '__main__':
    dict1 = {'wk1': [0, 0, 0, 0, 0, 0, 0], 'wk2': [0, 0, 0, 0, 0, 0, 0], 'wk3': [0, 0, 0, 0, 0, 0, 0], 'wk4': [0, 0, 0, 0, 0, 0, 0], 'wk5': [0, 0, 0, 0, 0, 0, 0], 'wk6': [1, 0, 0, 0, 0, 1, 1], 'wk7': [0, 0, 0, 0, 0, 0, 0], 'wk8': [0, 0, 0, 0, 0, 0, 0], 'wk9': [0, 0, 0, 0, 0, 0, 0], 'wk10': [5, 1, 0, 0, 0, 4, 1], 'wk11': [0, 0, 0, 0, 0, 0, 0], 'wk12': [0, 0, 0, 0, 0, 0, 0], 'wk13': [0, 0, 0, 0, 0, 0, 0], 'wk14': [0, 0, 0, 0, 0, 0, 0], 'wk15': [0, 0, 0, 0, 0, 0, 0], 'wk16': [0, 0, 0, 0, 0, 0, 0], 'wk17': [1, 0, 0, 0, 0, 1, 0], 'wk18': [0, 0, 0, 0, 0, 0, 0], 'wk19': [0, 0, 0, 0, 0, 0, 0], 'wk20': [0, 0, 0, 0, 0, 0, 0], 'wk21': [0, 0, 0, 0, 0, 0, 0], 'wk22': [0, 0, 0, 0, 0, 0, 0], 'wk23': [0, 0, 0, 0, 0, 0, 0], 'wk24': [0, 0, 0, 0, 0, 0, 0], 'wk25': [0, 0, 0, 0, 0, 0, 0], 'wk26': [0, 0, 0, 0, 0, 0, 0], 'wk27': [0, 0, 0, 0, 0, 0, 0], 'wk28': [0, 0, 0, 0, 0, 0, 0], 'wk29': [0, 0, 0, 0, 0, 0, 0], 'wk30': [0, 0, 0, 0, 0, 0, 0], 'wk31': [0, 0, 0, 0, 0, 0, 0], 'wk32': [2, 0, 0, 0, 0, 2, 0], 'wk33': [0, 0, 0, 0, 0, 0, 0], 'wk34': [0, 0, 0, 0, 0, 0, 0], 'wk35': [0, 0, 0, 0, 0, 0, 0], 'wk36': [0, 0, 0, 0, 0, 0, 0], 'wk37': [0, 0, 0, 0, 0, 0, 0], 'wk38': [0, 0, 0, 0, 0, 0, 0], 'wk39': [0, 0, 0, 0, 0, 0, 0], 'wk40': [0, 0, 0, 0, 0, 0, 0], 'wk41': [0, 0, 0, 0, 0, 0, 0], 'wk42': [0, 0, 0, 0, 0, 0, 0], 'wk43': [0, 0, 0, 0, 0, 0, 0], 'wk44': [0, 0, 0, 0, 0, 0, 0], 'wk45': [0, 0, 0, 0, 0, 0, 0], 'wk46': [0, 0, 0, 0, 0, 0, 0], 'wk47': [0, 0, 0, 0, 0, 0, 0], 'wk48': [0, 0, 0, 0, 0, 0, 0], 'wk49': [0, 0, 0, 0, 0, 0, 0], 'wk50': [0, 0, 0, 0, 0, 0, 0], 'wk51': [0, 0, 0, 0, 0, 0, 0], 'wk52': [0, 0, 0, 0, 0, 0, 0]}

    wb = openpyxl.Workbook()
    projectlist = ['E1 AL-LX110ST/AL-LW110ST', '中器S1','AA','BB','CC']
    for projectname in projectlist:
        index = projectlist.index(projectname)
        print(index)
        projectname = projectlist[index]
        if '/' in projectname:
            projectname = projectname.replace('/', '|')
        run_excel(wb,dict1, projectname, index)
